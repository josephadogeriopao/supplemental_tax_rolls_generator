import pandas as pd
import warnings
from dotenv import load_dotenv
import os
import sys
import shutil


def append_data_and_details(writer, df, data_sheet_name, details_sheet_name, source_file_path):
    """
    Writes data into the workbook with Excel SUM formulas for subtotals/totals.
    Fills empty numeric cells with 0 and styles headers/totals in bold.
    Returns a dictionary mapping (TaxYear, Column) -> exact string Excel cell coordinate of that Year's subtotal.
    """
    numeric_cols = [col for col in df.columns if pd.api.types.is_numeric_dtype(df[col])]
    exclude_cols = {'TAXYR', 'LUC', 'QUARTER'}
    cols_to_sum = [col for col in numeric_cols if col not in exclude_cols]
    num_format = '#,##0'

    df.head(0).to_excel(writer, sheet_name=data_sheet_name, index=False)

    workbook = writer.book
    worksheet = writer.sheets[data_sheet_name]

    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    bold_font = Font(bold=True)

    for col_idx in range(1, len(df.columns) + 1):
        worksheet.cell(row=1, column=col_idx).font = bold_font

    current_row = 2
    subtotal_rows = []
    subtotal_coordinates_by_year = {}

    for tax_year, group in df.groupby('TAXYR'):
        start_data_row = current_row

        for _, row in group.iterrows():
            for col_idx, col_name in enumerate(df.columns, start=1):
                val = row[col_name]
                cell = worksheet.cell(row=current_row, column=col_idx)

                if col_name in numeric_cols:
                    cell.value = int(val) if pd.notna(val) else 0
                    cell.number_format = num_format
                else:
                    cell.value = val
            current_row += 1

        end_data_row = current_row - 1

        worksheet.cell(row=current_row, column=1, value=f"{tax_year} Total").font = bold_font
        subtotal_rows.append(current_row)

        for col_idx, col_name in enumerate(df.columns, start=1):
            col_letter = get_column_letter(col_idx)
            if col_name in cols_to_sum:
                formula = f"=SUM({col_letter}{start_data_row}:{col_letter}{end_data_row})"
                cell = worksheet.cell(row=current_row, column=col_idx, value=formula)
                cell.font = bold_font
                cell.number_format = num_format

                # Track the individual subtotal row cell for global variables
                subtotal_coordinates_by_year[(int(tax_year), col_name)] = f"'{data_sheet_name}'!{col_letter}{current_row}"

            elif col_name == 'TAXYR':
                worksheet.cell(row=current_row, column=col_idx, value=tax_year).font = bold_font

        current_row += 2

    # Write Master Total Row
    worksheet.cell(row=current_row, column=1, value="Master Total").font = bold_font
    for col_idx, col_name in enumerate(df.columns, start=1):
        if col_name in cols_to_sum:
            col_letter = get_column_letter(col_idx)
            formula_cells = [f"{col_letter}{r}" for r in subtotal_rows]
            master_formula = f"=SUM({','.join(formula_cells)})" if formula_cells else "=0"

            cell = worksheet.cell(row=current_row, column=col_idx, value=master_formula)
            cell.font = bold_font
            cell.number_format = num_format

    # Handle Details Sheet Copy or Fallback Blank Sheet creation
    sql_copied = False
    try:
        xl_source = pd.ExcelFile(source_file_path)
        if len(xl_source.sheet_names) > 1:
            source_sql_sheet_name = xl_source.sheet_names[1]
            sql_df = pd.read_excel(source_file_path, sheet_name=source_sql_sheet_name, header=None)
            sql_df.to_excel(writer, sheet_name=details_sheet_name, index=False, header=False)
            print(f"   ℹ️ Copied incoming SQL sheet '{source_sql_sheet_name}' -> '{details_sheet_name}'")
            sql_copied = True
    except Exception as e:
        print(f"   ⚠️ Exception handled during source SQL tab copy: {e}")

    if not sql_copied:
        workbook.create_sheet(title=details_sheet_name)
        print(f"   ℹ️ Source SQL missing. Created an empty fallback tab -> '{details_sheet_name}'")

    return subtotal_coordinates_by_year