"""
DOCX Supplemental Tax Rolls Generator - Main application runner
------------------
Description: Main execution script that defines input/output configurations,
             writes structured details tabs, and overwrites worksheet-scoped
             local name manager variables directly with raw computed numeric sums
             instead of cross-sheet formulas.

Author: Joseph Adogeri
Version: 4.7.0
Since: 2026-08-03
File: main.py
License: MIT
"""
import pandas as pd
import warnings
from dotenv import load_dotenv
import os
import sys
import shutil
import openpyxl
from openpyxl.workbook.defined_name import DefinedName

# Import the engine module from your separate local processing file
from excel_engine import append_data_and_details
from utils.date import format_date
from utils.named_manager import get_tax_roll_initial_values

load_dotenv()

real_file = os.environ.get("REAL_XLSX_FILE", r"PATH TO PDF FILE")
pp_file = os.environ.get("PP_XLSX_FILE", r"PATH TO OUTPUT FOLDER/ DIRECTORY")
output_path = os.environ.get("OUTPUT_DIR", os.path.dirname(pp_file) if pp_file else "")


def main() -> None:
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
    if not all([real_file, pp_file, output_path]):
        print("❌ Error: Missing configuration paths or OUTPUT_DIR in your .env file!")
        return

    # 📌 USER ENTRY: Configure your active baseline target year and quarter
    TARGET_TAX_YEAR = 2026
    QUARTER = 2
    STR_DATE = format_date()

    # Construct the strict 4-year rolling target sequence (e.g., 2026, 2025, 2024, 2023)
    timeline_years = [TARGET_TAX_YEAR - i for i in range(4)]

    os.makedirs(output_path, exist_ok=True)

    script_dir = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(script_dir, "..", "..", "templates", "str_template.xlsx")
    final_output_path = os.path.join(output_path, "consolidated supplemental tax rolls.xlsx")

    if not os.path.exists(template_path):
        print(f"❌ Error: Base spreadsheet template not found at: {template_path}")
        return

    # Load source file data metrics
    real_df = pd.read_excel(real_file, sheet_name=0)
    pp_df = pd.read_excel(pp_file, sheet_name=0)

    if os.path.exists(final_output_path):
        try:
            with open(final_output_path, "r+"):
                pass
        except PermissionError:
            print(f"\n❌ ERROR: Permission Denied to file: {final_output_path}")
            print(f"👉 Please close the output spreadsheet in Microsoft Excel and rerun the script.")
            sys.exit(1)

    # 🚀 STEP 1: CLONE THE TEMPLATE DIRECTLY TO THE FINAL OUTPUT NAME
    print("--- STEP 1: CLONING BASE TEMPLATE WORKBOOK STRUCTURE ---")
    shutil.copyfile(template_path, final_output_path)

    # 🚀 STEP 2: WRITE COMPLEMENTARY PROPERTY SHEETS ONLY
    print("\n--- STEP 2: APPENDING LIVE DATA TABLES AND DETAIL SHEETS ---")
    with pd.ExcelWriter(final_output_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        print("Processing Real Property data...")
        real_subtotals = append_data_and_details(writer, real_df, "REAL PROPERTY DETAILS", "REAL PROPERTY SQL", real_file)

        print("Processing Real Data...", real_subtotals)

        print("Processing PP Property data...")
        pp_subtotals = append_data_and_details(writer, pp_df, "PERSONAL PROPERTY DETAILS", "PERSONAL PROPERTY SQL", pp_file)

        print("Processing PP Data...", pp_subtotals)

    # 🚀 STEP 3: UPDATE COPIED NAME VARIABLES DIRECTLY WITH STATIC VALUE TEXTS
    print("\n--- STEP 3: OVERWRITING LOCAL VARIABLE CONFIGURATIONS WITH STATIC NUMERIC SUMS ---")
    wb = openpyxl.load_workbook(final_output_path, data_only=False)

    target_sheet = "CONSOLIDATED SUMMARY"
    if target_sheet not in wb.sheetnames:
        print(f"❌ Error: Target sheet '{target_sheet}' missing from copied workbook structure!")
        return

    ws = wb[target_sheet]
    sheet_index = wb.sheetnames.index(target_sheet)
    real_diff_col = "TOTAL_ASMT_DIFF" if "TOTAL_ASMT_DIFF" in real_df.columns else "ASMT_TOTAL_DIFF"

    UPDATES = get_tax_roll_initial_values(QUARTER, TARGET_TAX_YEAR, STR_DATE)

    # 🔥 EVALUATE: Evaluate totals into static Python numbers. Do not map formulas starting with '='!
    # Instead of string formulas, we pass the raw dictionary value balance or fallback to default
    for index, year in enumerate(timeline_years, start=1):

        # 🏢 Real Estate Direct Balance Overwrite
        real_key = (year, real_diff_col)
        print(f"year: {year}, diff_col: {real_diff_col}, real_key: {real_key}")
        if real_key in real_subtotals:
            # Captures the raw computed sum from excel_engine instead of creating cell pointers
            print("value of real subtotal: ", real_subtotals[real_key])
            UPDATES[f"REAL_ESTATE_{index}"] = real_subtotals[real_key]

        # 📦 Personal Property Direct Balance Overwrite
        pp_key = (year, "NETASMT_DIFF")
        if pp_key in pp_subtotals:
            print("value of pp subtotal: ", pp_subtotals[pp_key])

            UPDATES[f"PERSONAL_PROPERTY_{index}"] = pp_subtotals[pp_key]

        # 🏡 Homestead Net Direct Balance Overwrite
        home_key = (year, "HOMESTEAD_DIFF")
        if home_key in real_subtotals:
            UPDATES[f"HOMESTEAD_EXEMPTION_NET_{index}"] = real_subtotals[home_key]

        print("updating data...",UPDATES)

    # 🔥 UNIFORM INJECTION PASS: All variables are written explicitly as clean static parameters
    for var_name, final_expression in UPDATES.items():
        new_dn = DefinedName(name=var_name, localSheetId=sheet_index)

        # FIX: Check if the parameter is a text string (like STR_DATE)
        if isinstance(final_expression, str):
            # Escape strings in literal double quotes to stop Excel's =@ array conversion engine bug
            new_dn.value = f'"{final_expression}"'
        else:
            # Numbers (ints/floats) remain raw string digit representations
            new_dn.value = str(final_expression)

        # Force commit changes into both local worksheet dictionary collections and global arrays
        ws.defined_names[var_name] = new_dn
        wb.defined_names.add(new_dn)
        print(f"   ✓ Name Manager: Hard Overwrite -> {var_name} = {new_dn.value}")

    # Enforce strict layout workbook sheet ordering
    desired_order = [target_sheet, "REAL PROPERTY DETAILS", "REAL PROPERTY SQL", "PERSONAL PROPERTY DETAILS", "PERSONAL PROPERTY SQL"]
    for target_idx, sheet_name in enumerate(desired_order):
        if sheet_name in wb.sheetnames:
            current_idx = wb.sheetnames.index(sheet_name)
            steps_to_move = target_idx - current_idx
            wb.move_sheet(wb[sheet_name], offset=steps_to_move)

    # Save and commit updates
    wb.save(final_output_path)
    wb.close()

    print(f"\n✅ Success! All local variables overwritten with calculated static sums directly.")


if __name__ == "__main__":
    main()
    sys.exit(0)

