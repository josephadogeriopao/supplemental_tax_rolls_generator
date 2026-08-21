"""
DOCX Supplemental Tax Rolls Generator - Class-Based Architecture
-----------------------------------------------------------------
Description: Class for name manager variable injection. Manages calculation updates,
             quote-escaping for text constants, variable injection, and sheet layout ordering.

Author: Joseph Adogeri
Version: 5.0.0
Since: 2026-08-21
File: name_manager_updater.py
License: MIT
"""

import os
import sys
import shutil
import warnings
from typing import List, Dict, Tuple, Any
import pandas as pd
import openpyxl
from openpyxl.workbook.defined_name import DefinedName
from dotenv import load_dotenv

# Pipeline engine imports
from excel_engine import append_data_and_details
from utils.date import format_date
from utils.named_manager import get_tax_roll_initial_values

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


class NameManagerUpdater:
    """
    Responsible for computing calculations, resolving rolling multi-year timelines,
    and overwriting spreadsheet-scoped local variables directly inside the Excel Name Manager.
    """

    def __init__(self, output_path: str, target_sheet: str = "CONSOLIDATED SUMMARY"):
        """
        Initializes the updater with the generated file path and destination sheet targets.

        Args:
            output_path (str): File system path to the target destination spreadsheet.
            target_sheet (str): The specific worksheet name holding the local variables.
                                Defaults to "CONSOLIDATED SUMMARY".
        """
        self.output_path = output_path
        self.target_sheet = target_sheet

    def update_variables(self, config: Any, real_subs: Dict, pp_subs: Dict, real_col: str) -> None:
        """
        Maps subtotal data matrix balances into variables and forces structural spreadsheet injection.

        Iterates through rolling timeline sequences, updates default lookup values with actual
        computed numbers, transforms text arguments into double-quoted safe values, and saves the openpyxl object.

        Args:
            config (ConfigManager): The active environmental configuration state object.
            real_subs (Dict): Real Property subtotal metrics grouped by year and target field keys.
            pp_subs (Dict): Personal Property subtotal metrics grouped by year and target field keys.
            real_col (str): The column text key name utilized to extract real assessment metrics.
        """
        print("\n--- STEP 3: OVERWRITING LOCAL VARIABLE CONFIGURATIONS WITH STATIC NUMERIC SUMS ---")
        wb = openpyxl.load_workbook(self.output_path, data_only=False)

        if self.target_sheet not in wb.sheetnames:
            print(f"❌ Error: Target sheet '{self.target_sheet}' missing from workbook structure!")
            return

        ws = wb[self.target_sheet]
        sheet_index = wb.sheetnames.index(self.target_sheet)

        # Build raw defaults dictionary
        updates = get_tax_roll_initial_values(config.quarter, config.target_tax_year, config.str_date)

        # Map dynamic sums into calculation update stack
        for index, year in enumerate(config.timeline_years, start=1):
            # Real Estate Overwrite
            real_key = (year, real_col)
            if real_key in real_subs:
                updates[f"REAL_ESTATE_{index}"] = real_subs[real_key]

            # Personal Property Overwrite
            pp_key = (year, "NETASMT_DIFF")
            if pp_key in pp_subs:
                updates[f"PERSONAL_PROPERTY_{index}"] = pp_subs[pp_key]

            # Homestead Net Overwrite
            home_key = (year, "HOMESTEAD_DIFF")
            if home_key in real_subs:
                updates[f"HOMESTEAD_EXEMPTION_NET_{index}"] = real_subs[home_key]

        # Inject metrics into spreadsheet via openpyxl
        for var_name, final_expression in updates.items():
            new_dn = DefinedName(name=var_name, localSheetId=sheet_index)

            if isinstance(final_expression, str):
                new_dn.value = f'"{final_expression}"'
            else:
                new_dn.value = str(final_expression)

            ws.defined_names[var_name] = new_dn
            wb.defined_names.add(new_dn)
            print(f"   ✓ Name Manager: Hard Overwrite -> {var_name} = {new_dn.value}")

        self._reorder_sheets(wb)
        wb.save(self.output_path)
        wb.close()

    def _reorder_sheets(self, wb: openpyxl.Workbook) -> None:
        """
        Enforces clean structural placement layout for sheets inside the final workbook.

        Internal helper function that re-orders worksheet index positions sequentially
        to guarantee summary parameters stay upfront.

        Args:
            wb (openpyxl.Workbook): The active, open workbook reference instance to manipulate.
        """
        desired_order = [self.target_sheet, "REAL PROPERTY DETAILS", "REAL PROPERTY SQL", "PERSONAL PROPERTY DETAILS",
                         "PERSONAL PROPERTY SQL"]
        for target_idx, sheet_name in enumerate(desired_order):
            if sheet_name in wb.sheetnames:
                current_idx = wb.sheetnames.index(sheet_name)
                wb.move_sheet(wb[sheet_name], offset=(target_idx - current_idx))
