import openpyxl
from openpyxl.workbook.defined_name import DefinedName

TEMPLATE_PATH = r"C:\Users\joseph.adogeri\Desktop\test_app\STR\str.xlsx"
OUTPUT_PATH = r"C:\Users\joseph.adogeri\Desktop\test_app\STR\str.xlsx"

UPDATES = {
    "HOMESTEAD_EXEMPTION_NET_1": 10000,
    "HOMESTEAD_EXEMPTION_NET_2": 20000,
    "HOMESTEAD_EXEMPTION_NET_3": 30000,
    "HOMESTEAD_EXEMPTION_NET_4": 40000,
    "PERSONAL_PROPERTY_1": 20220,
    "PERSONAL_PROPERTY_2": 20230,
    "PERSONAL_PROPERTY_3": 20240,
    "PERSONAL_PROPERTY_4": 20250,
    "QUARTER": 1,
    "REAL_ESTATE_1": 60000,
    "REAL_ESTATE_2": 70000,
    "REAL_ESTATE_3": 80000,
    "REAL_ESTATE_4": 90000,
    "TAX_YEAR": 2026,
    "STR_DATE": "APRIL 1, 2025"
}


def run():
    wb = openpyxl.load_workbook(TEMPLATE_PATH, data_only=False)

    target_sheet = "CONSOLIDATED SUMMARY"
    if target_sheet not in wb.sheetnames:
        print(f"Error: Sheet '{target_sheet}' not found!")
        return

    ws = wb[target_sheet]
    sheet_index = wb.sheetnames.index(target_sheet)

    print("--- STEP 1: CURRENT WORKSHEET VARIABLES ---")
    if hasattr(ws, "defined_names") and ws.defined_names:
        for var_name, dn in list(ws.defined_names.items()):
            print(f"Existing Name: {var_name} | Current Value: {dn.value}")
    else:
        print("No local variables currently visible on this sheet initialization.")

    print("\n--- STEP 2: UPDATING LOCAL SHEET VARIABLES ---")
    for var_name, new_value in UPDATES.items():
        # FIX: Pass the name first, and assign the clean text value directly to avoid the initialization error
        new_dn = DefinedName(name=var_name, localSheetId=sheet_index)
        new_dn.value = str(new_value)  # Pure string number without leading '=' prevents '@'

        # Inject/overwrite into the local sheet and global collections
        ws.defined_names[var_name] = new_dn
        wb.defined_names.add(new_dn)

        print(f"Updated: {var_name} -> {new_value}")

    wb.save(OUTPUT_PATH)
    print(f"\nWorkbook saved successfully to: {OUTPUT_PATH}")


if __name__ == "__main__":
    run()
